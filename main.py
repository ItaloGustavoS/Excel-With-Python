import openpyxl
import os

ws = openpyxl.load_workbook(
    "C:\\Users\\Felipe LM\\Documents\\GitHub\\Excel-With-Python\\Excel Files\\example.xlsx"
)
sheet_names = ws.sheetnames  # Correção: acessar os nomes das planilhas diretamente

for sheet_name in sheet_names:  # Itera sobre os nomes das planilhas
    print("Nome da planilha:", sheet_name)

    sheet = ws[sheet_name]  # Obtém a planilha pelo nome

    for row in sheet.iter_rows(values_only=True):  # Itera sobre as linhas da planilha
        for value in row:  # Itera sobre os valores de cada linha
            print(value)

    print()  # Adiciona uma linha em branco para separar as planilhas
